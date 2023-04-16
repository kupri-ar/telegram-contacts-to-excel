import datetime
import os

from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook


def create_contacts_excel_file(data):
    print('\nSaving all contacts to Excel file...')
    workbook = Workbook()
    sheet = workbook.active

    # write the headers to the first row
    headers = list(data[0].keys())
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col).value = header

    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 22
    sheet.column_dimensions['F'].width = 14

    # write the data to the remaining rows
    for row, item in enumerate(data, start=2):
        for col, header in enumerate(headers, start=1):
            if header != 'img':
                sheet.cell(row=row, column=col).value = item.get(header)
            else:
                img = item.get(header)
                if img:
                    sheet.row_dimensions[row].height = 70

                    cell = sheet.cell(row=row, column=col)
                    sheet.add_image(img, cell.coordinate)
                else:
                    sheet.row_dimensions[row].height = 30

            sheet.cell(row=row, column=col).alignment = Alignment(vertical='center')

    dirname = '../../xlsx'
    filename = 'TelegramContacts-{}.xlsx'.format(datetime.datetime.now().strftime("%Y-%m-%d %H %M %S"))

    # Check if the directory exists
    if not os.path.exists(dirname):
        os.makedirs(dirname)

    # save the workbook to a file
    workbook.save(filename='{}/{}'.format(dirname, filename))

    print('Contacts have been saved! \n'
          'You can find the Excel file with your Telegram contacts in the "xlsx" folder in the app directory.')
    print('Your file name: {}'.format(filename))
